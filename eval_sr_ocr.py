"""
Evaluate OCR accuracy on:
  1) Bicubic-upsampled LR images
  2) Super-Resolution model outputs

Metrics: character-level accuracy using Levenshtein distance.
Includes confusion matrix generation for OCR character predictions.
"""

import os
import glob
from pathlib import Path
import xml.etree.ElementTree as ET
from collections import defaultdict
import re

from PIL import Image
import Levenshtein
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.colors import LinearSegmentedColormap

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Excel export will be skipped.")
    print("Install with: pip install openpyxl")

from configs import SuperResolutionConfig, OCRConfig
from modules.super_resolution import SuperResolutionInference
from modules.ocr import OCRRecognizer


# Define multi-character tokens that should be treated as single units
MULTI_CHAR_TOKENS = ["Ein", "Gh", "Sad", "Sin", "Ta", "Zh"]

# Background token for confusion matrix
BG_TOKEN = "<BG>"


def parse_merge_pairs(merge_args: list) -> list:
    """
    Parse merge pair arguments into a list of (source, target) tuples.
    
    Args:
        merge_args: List of strings in format "SOURCE:TARGET"
                    e.g., ["3:2", "Sad:Sin"]
    
    Returns:
        List of (source, target) tuples.
        e.g., [("3", "2"), ("Sad", "Sin")]
    """
    pairs = []
    if not merge_args:
        return pairs
    
    for pair_str in merge_args:
        if ":" not in pair_str:
            print(f"Warning: Invalid merge pair format '{pair_str}'. Expected 'SOURCE:TARGET'. Skipping.")
            continue
        
        parts = pair_str.split(":", 1)
        source = parts[0].strip()
        target = parts[1].strip()
        
        if not source or not target:
            print(f"Warning: Empty source or target in merge pair '{pair_str}'. Skipping.")
            continue
        
        pairs.append((source, target))
    
    return pairs


def normalize_token(token: str, merge_pairs: list) -> str:
    """
    Normalize a single token by applying merge pairs.
    Each merge pair (source, target) replaces source token with target token.
    
    Args:
        token: The token to normalize.
        merge_pairs: List of (source, target) tuples.
    
    Returns:
        Normalized token.
    """
    if token is None:
        return None
    if token == BG_TOKEN:
        return token
    
    for source, target in merge_pairs:
        if token == source:
            return target
    
    return token


def normalize_tokens(tokens: list, merge_pairs: list) -> list:
    """
    Normalize a list of tokens by applying merge pairs.
    
    Args:
        tokens: List of tokens to normalize.
        merge_pairs: List of (source, target) tuples.
    
    Returns:
        List of normalized tokens.
    """
    if not tokens:
        return []
    return [normalize_token(t, merge_pairs) for t in tokens]


def normalize_string(text: str, merge_pairs: list) -> str:
    """
    Normalize a string by applying merge pairs.
    For multi-character tokens, tokenize first, normalize, then rejoin.
    
    Args:
        text: The string to normalize.
        merge_pairs: List of (source, target) tuples.
    
    Returns:
        Normalized string.
    """
    if text is None:
        return None
    if not merge_pairs:
        return text
    
    tokens = tokenize_plate_string(text)
    normalized = normalize_tokens(tokens, merge_pairs)
    return "".join(normalized)


def tokenize_plate_string(text: str) -> list:
    """
    Tokenize a plate string into individual tokens.
    Multi-character class names (Ein, Gh, Sad, Sin, Ta, Zh) are treated as single tokens.
    
    Args:
        text: The plate string (e.g., "12Ein45Gh7")
    
    Returns:
        List of tokens (e.g., ["1", "2", "Ein", "4", "5", "Gh", "7"])
    """
    if not text:
        return []
    
    tokens = []
    i = 0
    
    while i < len(text):
        matched = False
        # Check for multi-character tokens (longest match first)
        for token in sorted(MULTI_CHAR_TOKENS, key=len, reverse=True):
            if text[i:].startswith(token):
                tokens.append(token)
                i += len(token)
                matched = True
                break
        
        if not matched:
            # Single character token
            tokens.append(text[i])
            i += 1
    
    return tokens


def parse_xml_to_string(xml_path: Path, class_mapping) -> str | None:
    """
    Parses a PASCAL VOC XML file, sorts characters by their horizontal position,
    and returns the corresponding license plate string (mapped via class_mapping).
    """
    if not xml_path.exists():
        return None

    tree = ET.parse(str(xml_path))
    root = tree.getroot()

    characters = []
    for obj in root.findall("object"):
        class_name = obj.find("name").text
        xmin = int(obj.find("bndbox/xmin").text)

        mapped_name = class_mapping.get(class_name)
        if mapped_name is None:
            print(f"Warning: Class '{class_name}' in {xml_path} not in CLASS_MAPPING. Skipping this char.")
            continue

        characters.append((xmin, mapped_name))

    if not characters:
        return ""

    characters.sort(key=lambda x: x[0])
    return "".join([c[1] for c in characters])


def parse_xml_to_tokens(xml_path: Path, class_mapping) -> list | None:
    """
    Parses a PASCAL VOC XML file, sorts characters by their horizontal position,
    and returns the corresponding license plate as a list of tokens.
    """
    if not xml_path.exists():
        return None

    tree = ET.parse(str(xml_path))
    root = tree.getroot()

    characters = []
    for obj in root.findall("object"):
        class_name = obj.find("name").text
        xmin = int(obj.find("bndbox/xmin").text)

        mapped_name = class_mapping.get(class_name)
        if mapped_name is None:
            print(f"Warning: Class '{class_name}' in {xml_path} not in CLASS_MAPPING. Skipping this char.")
            continue

        characters.append((xmin, mapped_name))

    if not characters:
        return []

    characters.sort(key=lambda x: x[0])
    return [c[1] for c in characters]


def calculate_character_accuracy(gt_string: str, pred_string: str) -> float:
    """
    Character-level accuracy using Levenshtein distance:
    Accuracy = (len(GT) - distance) / len(GT)
    """
    if not gt_string:
        return 0.0 if pred_string else 1.0

    distance = Levenshtein.distance(gt_string, pred_string)
    acc = (len(gt_string) - distance) / len(gt_string)
    return max(0.0, acc)


def align_tokens_for_confusion(gt_tokens: list, pred_tokens: list, return_operations: bool = False):
    """
    Align ground truth and predicted token lists using Needleman-Wunsch algorithm.
    This properly handles insertions, deletions, and substitutions.
    
    Returns list of (gt_token, pred_token, operation) tuples where operation is:
    - 'match': tokens are identical
    - 'substitution': tokens differ
    - 'deletion': gt token has no corresponding pred token (gt=char, pred=background)
    - 'insertion': pred token has no corresponding gt token (gt=background, pred=char)
    """
    if not gt_tokens and not pred_tokens:
        return []
    
    if not gt_tokens:
        if return_operations:
            return [(BG_TOKEN, p, "insertion") for p in pred_tokens]
        return [(BG_TOKEN, p) for p in pred_tokens]
    
    if not pred_tokens:
        if return_operations:
            return [(g, BG_TOKEN, "deletion") for g in gt_tokens]
        return [(g, BG_TOKEN) for g in gt_tokens]
    
    m, n = len(gt_tokens), len(pred_tokens)
    
    gap = -1
    match_score = 2
    mismatch = -1
    
    score = [[0] * (n + 1) for _ in range(m + 1)]
    traceback = [[0] * (n + 1) for _ in range(m + 1)]
    
    for i in range(m + 1):
        score[i][0] = i * gap
        traceback[i][0] = 1
    for j in range(n + 1):
        score[0][j] = j * gap
        traceback[0][j] = 2
    
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if gt_tokens[i-1] == pred_tokens[j-1]:
                diag = score[i-1][j-1] + match_score
            else:
                diag = score[i-1][j-1] + mismatch
            
            up = score[i-1][j] + gap
            left = score[i][j-1] + gap
            
            max_score = max(diag, up, left)
            score[i][j] = max_score
            
            if max_score == diag:
                traceback[i][j] = 0
            elif max_score == up:
                traceback[i][j] = 1
            else:
                traceback[i][j] = 2
    
    aligned_pairs = []
    i, j = m, n
    
    while i > 0 or j > 0:
        if i > 0 and j > 0 and traceback[i][j] == 0:
            gt_tok = gt_tokens[i-1]
            pred_tok = pred_tokens[j-1]
            op = "match" if gt_tok == pred_tok else "substitution"
            if return_operations:
                aligned_pairs.append((gt_tok, pred_tok, op))
            else:
                aligned_pairs.append((gt_tok, pred_tok))
            i -= 1
            j -= 1
        elif i > 0 and (j == 0 or traceback[i][j] == 1):
            if return_operations:
                aligned_pairs.append((gt_tokens[i-1], BG_TOKEN, "deletion"))
            else:
                aligned_pairs.append((gt_tokens[i-1], BG_TOKEN))
            i -= 1
        else:
            if return_operations:
                aligned_pairs.append((BG_TOKEN, pred_tokens[j-1], "insertion"))
            else:
                aligned_pairs.append((BG_TOKEN, pred_tokens[j-1]))
            j -= 1
    
    aligned_pairs.reverse()
    return aligned_pairs


def print_alignment_details(gt_tokens: list, pred_tokens: list, label: str = ""):
    """Print detailed alignment information for debugging."""
    alignments = align_tokens_for_confusion(gt_tokens, pred_tokens, return_operations=True)
    
    print(f"\n  Alignment details for {label}:")
    print(f"  {'GT Token':<12} {'Pred Token':<12} {'Operation':<15}")
    print(f"  {'-'*12} {'-'*12} {'-'*15}")
    
    for item in alignments:
        gt_tok, pred_tok, op = item
        gt_display = gt_tok if gt_tok != BG_TOKEN else "(background)"
        pred_display = pred_tok if pred_tok != BG_TOKEN else "(background)"
        
        if op == "match":
            symbol = "✓"
        elif op == "substitution":
            symbol = "✗ (sub)"
        elif op == "deletion":
            symbol = "✗ (del) - missed character"
        else:
            symbol = "✗ (ins) - false positive"
        
        print(f"  {gt_display:<12} {pred_display:<12} {symbol}")
    
    return alignments


def create_white_green_cmap():
    """Create a custom colormap from white to green."""
    colors = ['#FFFFFF', '#E8F5E9', '#C8E6C9', '#A5D6A7', '#81C784', 
              '#66BB6A', '#4CAF50', '#43A047', '#388E3C', '#2E7D32', '#1B5E20']
    return LinearSegmentedColormap.from_list('WhiteGreen', colors, N=256)


class ConfusionMatrixBuilder:
    """Build and visualize confusion matrix for OCR tokens (characters/classes)."""
    
    def __init__(self, name: str = "OCR"):
        self.name = name
        self.confusion_counts = defaultdict(lambda: defaultdict(int))
        self.all_gt_tokens = set()
        self.all_pred_tokens = set()
        self.total_insertions = 0
        self.total_deletions = 0
    
    def add_prediction(self, gt_tokens: list, pred_tokens: list):
        """Add a prediction pair to the confusion matrix using token lists.
        
        IMPORTANT: Tokens should already be normalized before calling this method.
        """
        if gt_tokens is None:
            gt_tokens = []
        if pred_tokens is None:
            pred_tokens = []
        
        alignments = align_tokens_for_confusion(gt_tokens, pred_tokens, return_operations=True)
        
        for item in alignments:
            gt_token, pred_token, operation = item
            
            if operation == "match":
                self.confusion_counts[gt_token][gt_token] += 1
                self.all_gt_tokens.add(gt_token)
                self.all_pred_tokens.add(gt_token)
                
            elif operation == "substitution":
                self.confusion_counts[gt_token][pred_token] += 1
                self.all_gt_tokens.add(gt_token)
                self.all_pred_tokens.add(pred_token)
                
            elif operation == "deletion":
                self.confusion_counts[gt_token][BG_TOKEN] += 1
                self.all_gt_tokens.add(gt_token)
                self.all_pred_tokens.add(BG_TOKEN)
                self.total_deletions += 1
                
            elif operation == "insertion":
                self.confusion_counts[BG_TOKEN][pred_token] += 1
                self.all_gt_tokens.add(BG_TOKEN)
                self.all_pred_tokens.add(pred_token)
                self.total_insertions += 1
    
    def _sort_tokens(self, tokens):
        """Sort tokens: digits first (0-9), then single letters (A-Z), then multi-char tokens, then <BG>."""
        def sort_key(x):
            if x == BG_TOKEN:
                return (4, x)
            elif x.isdigit():
                return (0, int(x))
            elif len(x) == 1 and x.isalpha():
                return (1, x)
            else:
                return (2, x)
        return sorted(tokens, key=sort_key)
    
    def build_matrix(self):
        """Build the confusion matrix as numpy array."""
        all_tokens = self.all_gt_tokens | self.all_pred_tokens
        sorted_tokens = self._sort_tokens(all_tokens)
        
        n = len(sorted_tokens)
        matrix = np.zeros((n, n), dtype=int)
        
        for i, gt_token in enumerate(sorted_tokens):
            for j, pred_token in enumerate(sorted_tokens):
                matrix[i, j] = self.confusion_counts[gt_token][pred_token]
        
        return matrix, sorted_tokens, sorted_tokens
    
    def build_normalized_matrix(self):
        """Build the normalized confusion matrix (row-wise normalization)."""
        matrix, gt_tokens, pred_tokens = self.build_matrix()
        
        row_sums = matrix.sum(axis=1, keepdims=True)
        row_sums = np.where(row_sums == 0, 1, row_sums)
        normalized_matrix = matrix.astype(float) / row_sums
        
        return normalized_matrix, gt_tokens, pred_tokens
    
    def plot_and_save(self, save_path: str, figsize: tuple = None, normalized: bool = False):
        """Plot confusion matrix and save as image."""
        if normalized:
            matrix, gt_labels, pred_labels = self.build_normalized_matrix()
            fmt = '.2f'
            title_suffix = "(Normalized)"
            cbar_label = 'Proportion'
            vmin, vmax = 0, 1
        else:
            matrix, gt_labels, pred_labels = self.build_matrix()
            fmt = 'd'
            title_suffix = "(Count)"
            cbar_label = 'Count'
            vmin, vmax = None, None
        
        if len(gt_labels) == 0:
            print(f"Warning: No data to plot for {self.name} confusion matrix")
            return
        
        if figsize is None:
            n_labels = len(gt_labels)
            figsize = (max(14, n_labels * 0.6), max(10, n_labels * 0.5))
        
        fig, ax = plt.subplots(figsize=figsize)
        cmap = create_white_green_cmap()
        
        sns.heatmap(
            matrix,
            annot=True,
            fmt=fmt,
            cmap=cmap,
            xticklabels=pred_labels,
            yticklabels=gt_labels,
            ax=ax,
            cbar_kws={'label': cbar_label},
            vmin=vmin,
            vmax=vmax,
            linewidths=0.5,
            linecolor='lightgray'
        )
        
        ax.set_xlabel('Predicted Token', fontsize=12)
        ax.set_ylabel('Ground Truth Token', fontsize=12)
        ax.set_title(f'Confusion Matrix - {self.name} {title_suffix}', fontsize=14)
        
        plt.xticks(rotation=45, ha='right', fontsize=9)
        plt.yticks(rotation=0, fontsize=9)
        
        if BG_TOKEN in gt_labels:
            bg_idx = gt_labels.index(BG_TOKEN)
            ax.axhline(y=bg_idx, color='orange', linewidth=2)
            ax.axhline(y=bg_idx + 1, color='orange', linewidth=2)
            ax.axvline(x=bg_idx, color='orange', linewidth=2)
            ax.axvline(x=bg_idx + 1, color='orange', linewidth=2)
        
        plt.tight_layout()
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        plt.close()
        
        print(f"Confusion matrix saved: {save_path}")
    
    def save_to_excel(self, save_path: str):
        """Save both count and normalized confusion matrices to an Excel file."""
        if not HAS_OPENPYXL:
            print(f"Warning: openpyxl not installed. Skipping Excel export for {self.name}")
            return
        
        wb = openpyxl.Workbook()
        
        ws_count = wb.active
        ws_count.title = "Count"
        matrix_count, labels_count, _ = self.build_matrix()
        self._write_matrix_to_sheet(ws_count, matrix_count, labels_count, labels_count, 
                                     is_normalized=False)
        
        ws_norm = wb.create_sheet(title="Normalized")
        matrix_norm, labels_norm, _ = self.build_normalized_matrix()
        self._write_matrix_to_sheet(ws_norm, matrix_norm, labels_norm, labels_norm, 
                                     is_normalized=True)
        
        ws_summary = wb.create_sheet(title="Summary")
        self._write_summary_to_sheet(ws_summary, matrix_count, labels_count)
        
        wb.save(save_path)
        print(f"Confusion matrix Excel saved: {save_path}")
    
    def _write_matrix_to_sheet(self, ws, matrix, gt_labels, pred_labels, is_normalized=False):
        """Write a confusion matrix to an Excel worksheet with formatting."""
        n = len(gt_labels)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        bg_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        diag_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        title = f"Confusion Matrix - {self.name} ({'Normalized' if is_normalized else 'Count'})"
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(n + 3, 20))
        
        if is_normalized:
            ws.cell(row=2, column=1, value="Note: Values are row-normalized (each row sums to 1.0)").font = Font(italic=True, size=10)
        else:
            ws.cell(row=2, column=1, value="Note: Values are raw counts").font = Font(italic=True, size=10)
        
        start_row = 4
        
        cell = ws.cell(row=start_row, column=1, value="GT \\ Pred")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = center_align
        cell.border = thin_border
        
        for j, pred_label in enumerate(pred_labels):
            cell = ws.cell(row=start_row, column=j + 2, value=pred_label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            if pred_label == BG_TOKEN:
                cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        
        cell = ws.cell(row=start_row, column=n + 2, value="Total")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = center_align
        cell.border = thin_border
        
        if not is_normalized:
            cell = ws.cell(row=start_row, column=n + 3, value="Accuracy")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = center_align
            cell.border = thin_border
        
        for i, gt_label in enumerate(gt_labels):
            row = start_row + 1 + i
            
            cell = ws.cell(row=row, column=1, value=gt_label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            if gt_label == BG_TOKEN:
                cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
            
            row_sum = 0
            for j in range(n):
                value = matrix[i, j]
                cell = ws.cell(row=row, column=j + 2)
                
                if is_normalized:
                    cell.value = round(float(value), 4)
                    cell.number_format = '0.00%'
                else:
                    cell.value = int(value)
                    row_sum += int(value)
                
                cell.alignment = center_align
                cell.border = thin_border
                
                if i == j:
                    cell.fill = diag_fill
                    cell.font = Font(bold=True)
                
                if gt_label == BG_TOKEN or pred_labels[j] == BG_TOKEN:
                    if i != j:
                        cell.fill = bg_fill
                
                if i != j and ((is_normalized and value > 0) or (not is_normalized and value > 0)):
                    if gt_label != BG_TOKEN and pred_labels[j] != BG_TOKEN:
                        cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            
            if not is_normalized:
                cell = ws.cell(row=row, column=n + 2, value=row_sum)
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = Font(bold=True)
                
                correct = int(matrix[i, i])
                acc = correct / row_sum if row_sum > 0 else 0
                cell = ws.cell(row=row, column=n + 3)
                if gt_label == BG_TOKEN:
                    cell.value = "N/A"
                    cell.number_format = '@'
                else:
                    cell.value = acc
                    cell.number_format = '0.00%'
                cell.alignment = center_align
                cell.border = thin_border
            else:
                row_total = float(matrix[i, :].sum())
                cell = ws.cell(row=row, column=n + 2, value=round(row_total, 4))
                cell.number_format = '0.00'
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = Font(bold=True)
        
        if not is_normalized:
            total_row = start_row + 1 + n
            cell = ws.cell(row=total_row, column=1, value="Total")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = center_align
            cell.border = thin_border
            
            grand_total = 0
            for j in range(n):
                col_sum = int(matrix[:, j].sum())
                grand_total += col_sum
                cell = ws.cell(row=total_row, column=j + 2, value=col_sum)
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = Font(bold=True)
            
            cell = ws.cell(row=total_row, column=n + 2, value=grand_total)
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = Font(bold=True)
        
        for col_idx in range(1, n + 4):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(start_row, start_row + n + 3):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_length + 2, 8)
    
    def _write_summary_to_sheet(self, ws, matrix, labels):
        """Write summary statistics to an Excel worksheet."""
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=10)
        
        row = 1
        ws.cell(row=row, column=1, value=f"Summary - {self.name}").font = Font(bold=True, size=14)
        row += 2
        
        ws.cell(row=row, column=1, value="Per-Token Accuracy").font = header_font
        row += 1
        
        ws.cell(row=row, column=1, value="Token").font = subheader_font
        ws.cell(row=row, column=2, value="Correct").font = subheader_font
        ws.cell(row=row, column=3, value="Total").font = subheader_font
        ws.cell(row=row, column=4, value="Accuracy").font = subheader_font
        ws.cell(row=row, column=5, value="Main Confusions").font = subheader_font
        row += 1
        
        total_correct = 0
        total_count = 0
        
        for i, gt_token in enumerate(labels):
            row_sum = int(matrix[i, :].sum())
            if row_sum > 0:
                correct = int(matrix[i, i])
                accuracy = correct / row_sum
                
                if gt_token != BG_TOKEN:
                    total_correct += correct
                    total_count += row_sum
                
                confusions = []
                for j, pred_token in enumerate(labels):
                    if i != j and matrix[i, j] > 0:
                        confusions.append(f"{pred_token}:{int(matrix[i, j])}")
                
                ws.cell(row=row, column=1, value=gt_token)
                ws.cell(row=row, column=2, value=correct)
                ws.cell(row=row, column=3, value=row_sum)
                
                if gt_token == BG_TOKEN:
                    ws.cell(row=row, column=4, value="N/A")
                else:
                    cell = ws.cell(row=row, column=4, value=accuracy)
                    cell.number_format = '0.00%'
                
                ws.cell(row=row, column=5, value=", ".join(confusions) if confusions else "-")
                row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="Overall (excl. BG)").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_correct).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_count).font = Font(bold=True)
        if total_count > 0:
            cell = ws.cell(row=row, column=4, value=total_correct / total_count)
            cell.number_format = '0.00%'
            cell.font = Font(bold=True)
        
        row += 2
        ws.cell(row=row, column=1, value="Background Statistics").font = header_font
        row += 1
        
        if BG_TOKEN in labels:
            bg_idx = labels.index(BG_TOKEN)
            deletions = sum(int(matrix[i, bg_idx]) for i in range(len(labels)) if i != bg_idx)
            insertions = sum(int(matrix[bg_idx, j]) for j in range(len(labels)) if j != bg_idx)
            
            ws.cell(row=row, column=1, value="Deletions (char → background)")
            ws.cell(row=row, column=2, value=deletions)
            row += 1
            ws.cell(row=row, column=1, value="Insertions (background → char)")
            ws.cell(row=row, column=2, value=insertions)
        else:
            ws.cell(row=row, column=1, value="No background errors recorded")
        
        for col_idx in range(1, 6):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
    
    def print_summary(self):
        """Print summary statistics from confusion matrix."""
        matrix, labels, _ = self.build_matrix()
        
        if len(labels) == 0:
            print(f"No data for {self.name}")
            return
        
        print(f"\n{'='*60}")
        print(f"{self.name} - Per-token accuracy:")
        print(f"{'='*60}")
        
        total_correct = 0
        total_count = 0
        
        for i, gt_token in enumerate(labels):
            row_sum = matrix[i, :].sum()
            if row_sum > 0:
                correct = matrix[i, i]
                accuracy = correct / row_sum * 100
                
                if gt_token != BG_TOKEN:
                    total_correct += correct
                    total_count += row_sum
                
                confusions = []
                for j, pred_token in enumerate(labels):
                    if i != j and matrix[i, j] > 0:
                        confusions.append(f"{pred_token}:{matrix[i, j]}")
                
                conf_str = f" | Confused with: {', '.join(confusions)}" if confusions else ""
                
                if gt_token == BG_TOKEN:
                    print(f"  '{gt_token}' (Background -> False Positives): {row_sum} total{conf_str}")
                else:
                    print(f"  '{gt_token}': {correct}/{row_sum} ({accuracy:.1f}%){conf_str}")
        
        if total_count > 0:
            overall_acc = total_correct / total_count * 100
            print(f"\n  Overall (excluding background): {total_correct}/{total_count} ({overall_acc:.1f}%)")
        
        if BG_TOKEN in labels:
            bg_idx = labels.index(BG_TOKEN)
            deletions = sum(matrix[i, bg_idx] for i in range(len(labels)) if i != bg_idx)
            insertions = sum(matrix[bg_idx, j] for j in range(len(labels)) if j != bg_idx)
            
            print(f"\n  Background Statistics:")
            print(f"    Deletions (char -> background): {deletions}")
            print(f"    Insertions (background -> char): {insertions}")


def plot_combined_matrices(cms: list, output_dir: Path, normalized: bool = False):
    """Plot combined confusion matrices for all methods in a single figure."""
    n_matrices = len(cms)
    fig, axes = plt.subplots(1, n_matrices, figsize=(10 * n_matrices, 10))
    
    if n_matrices == 1:
        axes = [axes]
    
    cmap = create_white_green_cmap()
    
    for idx, (cm, title) in enumerate(cms):
        ax = axes[idx]
        
        if normalized:
            matrix, gt_labels, pred_labels = cm.build_normalized_matrix()
            fmt = '.2f'
            title_suffix = "(Normalized)"
            cbar_label = 'Proportion'
            vmin, vmax = 0, 1
        else:
            matrix, gt_labels, pred_labels = cm.build_matrix()
            fmt = 'd'
            title_suffix = "(Count)"
            cbar_label = 'Count'
            vmin, vmax = None, None
        
        if len(gt_labels) > 0:
            sns.heatmap(
                matrix,
                annot=True,
                fmt=fmt,
                cmap=cmap,
                xticklabels=pred_labels,
                yticklabels=gt_labels,
                ax=ax,
                cbar_kws={'label': cbar_label},
                vmin=vmin,
                vmax=vmax,
                linewidths=0.5,
                linecolor='lightgray'
            )
            ax.set_xlabel('Predicted Token', fontsize=10)
            ax.set_ylabel('Ground Truth Token', fontsize=10)
            ax.set_title(f'{title}\n{title_suffix}', fontsize=12)
            ax.tick_params(axis='x', rotation=45)
            
            if BG_TOKEN in gt_labels:
                bg_idx = gt_labels.index(BG_TOKEN)
                ax.axhline(y=bg_idx, color='orange', linewidth=2)
                ax.axhline(y=bg_idx + 1, color='orange', linewidth=2)
                ax.axvline(x=bg_idx, color='orange', linewidth=2)
                ax.axvline(x=bg_idx + 1, color='orange', linewidth=2)
        else:
            ax.text(0.5, 0.5, 'No Data', ha='center', va='center', fontsize=14)
            ax.set_title(f'{title}\n{title_suffix}', fontsize=12)
    
    plt.tight_layout()
    
    suffix = "normalized" if normalized else "count"
    save_path = output_dir / f"confusion_matrix_combined_{suffix}.png"
    plt.savefig(str(save_path), dpi=150, bbox_inches='tight')
    print(f"Combined confusion matrix ({suffix}) saved: {save_path}")
    
    return fig


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Evaluate OCR accuracy on bicubic vs Super-Resolution outputs"
    )
    parser.add_argument(
        "--lr-dir",
        type=str,
        required=True,
        help="Directory with low-resolution plate images",
    )
    parser.add_argument(
        "--hr-dir",
        type=str,
        required=True,
        help="Directory with high-resolution plate images",
    )
    parser.add_argument(
        "--xml-dir",
        type=str,
        required=True,
        help="Directory with PASCAL VOC XML annotations for plates",
    )
    parser.add_argument(
        "--ocr-model",
        type=str,
        required=True,
        help="Path to YOLOv8 OCR model (.pt)",
    )
    parser.add_argument(
        "--sr-model",
        type=str,
        required=True,
        help="Path to Super-Resolution model weights (.pth)",
    )
    parser.add_argument(
        "--scale-factor",
        type=int,
        default=8,
        help="Upscaling factor for bicubic and SR model (must match SR model)",
    )
    parser.add_argument(
        "--device",
        type=str,
        default="cuda",
        choices=["cuda", "cpu"],
        help="Device to use",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default="eval_results",
        help="Directory to save confusion matrix images (default: eval_results)",
    )
    parser.add_argument(
        "--show-alignment",
        action="store_true",
        help="Show detailed alignment for each image (verbose output)",
    )
    parser.add_argument(
        "--merge",
        type=str,
        nargs="*",
        default=[],
        help=(
            "Merge (normalize) token pairs in format SOURCE:TARGET. "
            "All occurrences of SOURCE will be treated as TARGET. "
            "Can specify multiple pairs. "
            "Example: --merge 3:2 Sad:Sin"
        ),
    )
    args = parser.parse_args()

    lr_dir = Path(args.lr_dir)
    hr_dir = Path(args.hr_dir)
    xml_dir = Path(args.xml_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # --- Parse merge pairs ---
    merge_pairs = parse_merge_pairs(args.merge)
    if merge_pairs:
        print("Token merge rules (for accuracy & confusion matrix):")
        for source, target in merge_pairs:
            print(f"  '{source}' → '{target}'")
        print()

    # --- Set up SR inference (force apply to all images) ---
    sr_config = SuperResolutionConfig(
        model_path=args.sr_model,
        device=args.device,
        debug=False,
        apply_threshold=False,
        scale_factor=args.scale_factor,
    )
    sr_infer = SuperResolutionInference(sr_config)

    # --- Set up OCR recognizer ---
    ocr_config = OCRConfig(
        model_path=args.ocr_model,
        device=args.device,
        debug=False,
    )
    ocr = OCRRecognizer(ocr_config)
    class_mapping = ocr_config.class_mapping

    # --- Collect LR images ---
    image_paths = []
    for ext in ("*.jpg", "*.jpeg", "*.png", "*.bmp", "*.tif", "*.tiff"):
        image_paths.extend(glob.glob(str(lr_dir / ext)))
    image_paths = sorted(image_paths)

    if not image_paths:
        print(f"Error: No images found in '{lr_dir}'.")
        return

    print(f"Found {len(image_paths)} LR images.")

    # Initialize confusion matrix builders
    cm_hr = ConfusionMatrixBuilder(name="HR (High Resolution)")
    cm_bicubic = ConfusionMatrixBuilder(name="Bicubic Upsampling")
    cm_sr = ConfusionMatrixBuilder(name="Super Resolution")

    bicubic_accuracies = []
    hr_accuracies = []
    sr_accuracies = []
    bicubic_perfect = 0
    hr_perfect = 0
    sr_perfect = 0

    for img_path in image_paths:
        img_path = Path(img_path)
        image_filename = img_path.name
        hr_filename = img_path.stem + ".jpg"
        hr_path = hr_dir / hr_filename
        xml_filename = img_path.stem + ".xml"
        xml_path = xml_dir / xml_filename

        # --- Ground truth string and tokens from XML ---
        gt_string = parse_xml_to_string(xml_path, class_mapping)
        gt_tokens = parse_xml_to_tokens(xml_path, class_mapping)
        
        if gt_string is None:
            print(f"Warning: No XML found for {image_filename}, skipping.")
            continue

        # --- Load LR image ---
        lr_img = Image.open(img_path).convert("RGB")
        w, h = lr_img.size

        # --- HR image OCR ---
        hr_img = Image.open(hr_path).convert("RGB")
        hr_pred, _ = ocr.recognize(hr_img)
        hr_pred_tokens = tokenize_plate_string(hr_pred) if hr_pred else []
        hr_acc = calculate_character_accuracy(gt_string, hr_pred)
        hr_accuracies.append(hr_acc)
        if hr_acc > 0.99:
            hr_perfect += 1

        # --- 1) Bicubic baseline upscaling ---
        bicubic_img = lr_img.resize(
            (w * args.scale_factor, h * args.scale_factor),
            resample=Image.BICUBIC,
        )

        # --- 2) SR model upscaling ---
        sr_img = sr_infer.enhance(lr_img)

        # --- OCR on bicubic ---
        bicubic_pred, _ = ocr.recognize(bicubic_img)
        bicubic_pred_tokens = tokenize_plate_string(bicubic_pred) if bicubic_pred else []

        # --- OCR on SR output ---
        sr_pred, _ = ocr.recognize(sr_img)
        sr_pred_tokens = tokenize_plate_string(sr_pred) if sr_pred else []

        # --- Normalize strings for accuracy calculation ---
        gt_string_normalized = normalize_string(gt_string, merge_pairs)
        hr_pred_normalized = normalize_string(hr_pred, merge_pairs)
        bicubic_pred_normalized = normalize_string(bicubic_pred, merge_pairs)
        sr_pred_normalized = normalize_string(sr_pred, merge_pairs)

        # --- Normalize tokens for confusion matrix ---
        gt_tokens_normalized = normalize_tokens(gt_tokens, merge_pairs)
        hr_pred_tokens_normalized = normalize_tokens(hr_pred_tokens, merge_pairs)
        bicubic_pred_tokens_normalized = normalize_tokens(bicubic_pred_tokens, merge_pairs)
        sr_pred_tokens_normalized = normalize_tokens(sr_pred_tokens, merge_pairs)

        # --- Add NORMALIZED tokens to confusion matrices ---
        cm_hr.add_prediction(gt_tokens_normalized, hr_pred_tokens_normalized)
        cm_bicubic.add_prediction(gt_tokens_normalized, bicubic_pred_tokens_normalized)
        cm_sr.add_prediction(gt_tokens_normalized, sr_pred_tokens_normalized)

        # --- Calculate accuracies using normalized strings ---
        bicubic_acc = calculate_character_accuracy(gt_string_normalized, bicubic_pred_normalized)
        bicubic_accuracies.append(bicubic_acc)
        if bicubic_acc > 0.99:
            bicubic_perfect += 1

        sr_acc = calculate_character_accuracy(gt_string_normalized, sr_pred_normalized)
        sr_accuracies.append(sr_acc)
        if sr_acc > 0.99:
            sr_perfect += 1

        print("-" * 60)
        print(f"Image: {image_filename}")
        print(f"  GT:         {gt_string} -> tokens: {gt_tokens} -> normalized: {gt_tokens_normalized}")
        print(f"  HR_Image:   {hr_pred} -> tokens: {hr_pred_tokens} -> normalized: {hr_pred_tokens_normalized}  (acc={hr_acc:.2%})")
        print(f"  Bicubic:    {bicubic_pred} -> tokens: {bicubic_pred_tokens} -> normalized: {bicubic_pred_tokens_normalized}  (acc={bicubic_acc:.2%})")
        print(f"  SR model:   {sr_pred} -> tokens: {sr_pred_tokens} -> normalized: {sr_pred_tokens_normalized}  (acc={sr_acc:.2%})")
        
        # Show detailed alignment if requested or if there are errors
        if args.show_alignment or (sr_acc < 0.99 and sr_acc > 0):
            print_alignment_details(gt_tokens_normalized, sr_pred_tokens_normalized, "SR model (normalized)")

    # --- Final summary ---
    if bicubic_accuracies:
        avg_hr = sum(hr_accuracies) / len(hr_accuracies)
        avg_bicubic = sum(bicubic_accuracies) / len(bicubic_accuracies)
        avg_sr = sum(sr_accuracies) / len(sr_accuracies)

        print("\n" + "=" * 60)
        print("EVALUATION COMPLETE")
        print("=" * 60)
        if merge_pairs:
            merge_desc = ", ".join([f"'{s}'→'{t}'" for s, t in merge_pairs])
            print(f"Token merges applied: {merge_desc}")
        print(f"Images evaluated: {len(bicubic_accuracies)}")
        print(f"Average char accuracy (HR): {avg_hr:.2%}")
        print(f"Average char accuracy (Bicubic): {avg_bicubic:.2%}")
        print(f"Average char accuracy (SR):      {avg_sr:.2%}")
        print(f"Perfect plates (acc>0.99) HR: {hr_perfect}")
        print(f"Perfect plates (acc>0.99) Bicubic: {bicubic_perfect}")
        print(f"Perfect plates (acc>0.99) SR:      {sr_perfect}")
        print("=" * 60)
        
        # --- Generate and save confusion matrices ---
        print("\n" + "=" * 60)
        print("GENERATING CONFUSION MATRICES")
        print("=" * 60)
        
        # HR
        cm_hr.plot_and_save(str(output_dir / "confusion_matrix_hr_count.png"), normalized=False)
        cm_hr.plot_and_save(str(output_dir / "confusion_matrix_hr_normalized.png"), normalized=True)
        cm_hr.save_to_excel(str(output_dir / "confusion_matrix_hr.xlsx"))
        
        # Bicubic
        cm_bicubic.plot_and_save(str(output_dir / "confusion_matrix_bicubic_count.png"), normalized=False)
        cm_bicubic.plot_and_save(str(output_dir / "confusion_matrix_bicubic_normalized.png"), normalized=True)
        cm_bicubic.save_to_excel(str(output_dir / "confusion_matrix_bicubic.xlsx"))
        
        # SR
        cm_sr.plot_and_save(str(output_dir / "confusion_matrix_sr_count.png"), normalized=False)
        cm_sr.plot_and_save(str(output_dir / "confusion_matrix_sr_normalized.png"), normalized=True)
        cm_sr.save_to_excel(str(output_dir / "confusion_matrix_sr.xlsx"))
        
        # Print summaries
        cm_hr.print_summary()
        cm_bicubic.print_summary()
        cm_sr.print_summary()
        
        # --- Combined figures ---
        print("\n" + "=" * 60)
        print("GENERATING COMBINED CONFUSION MATRICES")
        print("=" * 60)
        
        cms_list = [
            (cm_hr, "HR (High Resolution)"),
            (cm_bicubic, "Bicubic Upsampling"),
            (cm_sr, "Super Resolution")
        ]
        
        fig_count = plot_combined_matrices(cms_list, output_dir, normalized=False)
        fig_normalized = plot_combined_matrices(cms_list, output_dir, normalized=True)
        
        print("\n" + "=" * 60)
        print("DISPLAYING CONFUSION MATRICES")
        print("=" * 60)
        
        plt.show()
        
    else:
        print("\nNo images were evaluated.")
        

if __name__ == "__main__":
    main()